#!/usr/bin/env python3
"""
NESMA 功能点清单 Excel 生成器
生成标准格式的功能点评估报告
"""

import io
from datetime import datetime
from typing import Dict, List, Any, Optional


def generate_excel(function_data: Dict[str, Any], project_name: str = "软件项目") -> bytes:
    """
    生成功能点清单 Excel 文件
    
    Args:
        function_data: 功能点数据，包含 details 和 total
        project_name: 项目名称
        
    Returns:
        Excel 文件的字节内容
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "功能点清单"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    
    # 标题
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"{project_name} - NESMA 功能点评估清单"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.row_dimensions[1].height = 35
    
    # 日期和方法
    ws.merge_cells('A2:G2')
    method_desc = function_data.get('method_description', '')
    date_cell = ws['A2']
    date_cell.value = f"评估日期: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 计算方法: {method_desc}"
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    date_cell.font = Font(size=10)
    ws.row_dimensions[2].height = 25
    
    # 表头
    headers = ['序号', '功能模块', '功能描述', '功能类型', '复杂度', '功能点', '备注']
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    header_font = Font(bold=True, size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[4].height = 25
    
    # 数据行
    row = 5
    details = function_data.get('details', [])
    
    # 类型颜色映射
    type_colors = {
        'ILF': 'E2EFDA',  # 浅绿
        'EIF': 'FCE4D6',  # 浅橙
        'EI': 'DDEBF7',   # 浅蓝
        'EO': 'F4CCCC',   # 浅红
        'EQ': 'FFF2CC'    # 浅黄
    }
    
    for idx, item in enumerate(details, 1):
        fp_type = item.get('type', '')
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.get('module', ''))
        ws.cell(row=row, column=3, value=item.get('description', ''))
        ws.cell(row=row, column=4, value=fp_type)
        ws.cell(row=row, column=5, value=item.get('complexity', ''))
        ws.cell(row=row, column=6, value=item.get('fp', 0))
        ws.cell(row=row, column=7, value=item.get('note', ''))
        
        # 设置对齐
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # 添加类型颜色
            if col == 4 and fp_type in type_colors:
                cell.fill = PatternFill(start_color=type_colors[fp_type], 
                                       end_color=type_colors[fp_type], 
                                       fill_type='solid')
        
        # 序号和功能点居中
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[row].height = 22
        row += 1
    
    # 空行
    row += 1
    
    # 统计行
    ws.merge_cells(f'A{row}:E{row}')
    total_label = ws.cell(row=row, column=1)
    total_label.value = "功能点总计"
    total_label.font = Font(bold=True, size=12)
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    
    total_cell = ws.cell(row=row, column=6)
    total_cell.value = function_data.get('total', 0)
    total_cell.font = Font(bold=True, size=14, color='FFFFFF')
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    ws.row_dimensions[row].height = 30
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for row_cells in ws[f'A4:G{row}']:
        for cell in row_cells:
            cell.border = thin_border
    
    # 添加类型汇总表
    row += 3
    ws.merge_cells(f'A{row}:G{row}')
    summary_title = ws.cell(row=row, column=1)
    summary_title.value = "功能类型汇总"
    summary_title.font = Font(bold=True, size=12)
    summary_title.alignment = Alignment(horizontal='left', vertical='center')
    summary_title.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    ws.row_dimensions[row].height = 25
    
    row += 1
    
    # 汇总表头
    summary_headers = ['功能类型', '类型名称', '数量', '功能点数', '占比']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    row += 1
    
    type_names = {
        'ILF': '内部逻辑文件',
        'EIF': '外部接口文件',
        'EI': '外部输入',
        'EO': '外部输出',
        'EQ': '外部查询'
    }
    
    type_summary = function_data.get('type_summary', {})
    total_fp = function_data.get('total', 1)
    
    for fp_type in ['ILF', 'EIF', 'EI', 'EO', 'EQ']:
        summary = type_summary.get(fp_type, {'count': 0, 'fp': 0})
        ws.cell(row=row, column=1, value=fp_type)
        ws.cell(row=row, column=2, value=type_names.get(fp_type, ''))
        ws.cell(row=row, column=3, value=summary['count'])
        ws.cell(row=row, column=4, value=summary['fp'])
        
        percentage = (summary['fp'] / total_fp * 100) if total_fp > 0 else 0
        ws.cell(row=row, column=5, value=f"{percentage:.1f}%")
        
        # 居中对齐
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # 添加说明
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    note_cell = ws.cell(row=row, column=1)
    note_cell.value = (
        "说明：本清单基于 NESMA 功能点分析方法生成\n"
        "ILF(内部逻辑文件) - 系统内部维护的数据\n"
        "EIF(外部接口文件) - 系统引用但外部维护的数据\n"
        "EI(外部输入) - 用户或其他系统提供的数据\n"
        "EO(外部输出) - 经过复杂处理的数据输出\n"
        "EQ(外部查询) - 简单的数据查询和显示"
    )
    note_cell.font = Font(size=9)
    note_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 80
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_summary_excel(results: List[Dict[str, Any]], project_name: str = "项目汇总") -> bytes:
    """
    生成多个项目的汇总 Excel
    
    Args:
        results: 多个项目的分析结果
        project_name: 汇总项目名称
        
    Returns:
        Excel 文件的字节内容
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "项目汇总"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # 标题
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"{project_name} - NESMA 功能点汇总报告"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.row_dimensions[1].height = 35
    
    # 日期
    ws.merge_cells('A2:I2')
    date_cell = ws['A2']
    date_cell.value = f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 20
    
    # 表头
    headers = ['序号', '项目名称', 'ILF', 'EIF', 'EI', 'EO', 'EQ', '功能点总计', '计算方法']
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行
    row = 5
    grand_total = 0
    
    for idx, result in enumerate(results, 1):
        type_summary = result.get('type_summary', {})
        total = result.get('total', 0)
        grand_total += total
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=result.get('project_name', f'项目{idx}'))
        ws.cell(row=row, column=3, value=type_summary.get('ILF', {}).get('fp', 0))
        ws.cell(row=row, column=4, value=type_summary.get('EIF', {}).get('fp', 0))
        ws.cell(row=row, column=5, value=type_summary.get('EI', {}).get('fp', 0))
        ws.cell(row=row, column=6, value=type_summary.get('EO', {}).get('fp', 0))
        ws.cell(row=row, column=7, value=type_summary.get('EQ', {}).get('fp', 0))
        ws.cell(row=row, column=8, value=total)
        ws.cell(row=row, column=9, value=result.get('method', 'detailed'))
        
        # 居中对齐
        for col in range(1, 10):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # 总计行
    ws.merge_cells(f'A{row}:B{row}')
    total_label = ws.cell(row=row, column=1)
    total_label.value = "总计"
    total_label.font = Font(bold=True, size=12)
    total_label.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=row, column=8, value=grand_total)
    ws.cell(row=row, column=8).font = Font(bold=True, size=12)
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_cells in ws[f'A4:I{row}']:
        for cell in row_cells:
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


if __name__ == '__main__':
    # 测试示例
    test_data = {
        'total': 116,
        'method_description': '详细法 - 根据实际复杂度计算',
        'type_summary': {
            'ILF': {'count': 3, 'fp': 30},
            'EIF': {'count': 1, 'fp': 7},
            'EI': {'count': 2, 'fp': 8},
            'EO': {'count': 2, 'fp': 12},
            'EQ': {'count': 3, 'fp': 12}
        },
        'details': [
            {'id': '1', 'module': '商品管理', 'description': '商品表', 'type': 'ILF', 'complexity': '中', 'fp': 10},
            {'id': '2', 'module': '商品管理', 'description': '分类表', 'type': 'ILF', 'complexity': '低', 'fp': 7},
            {'id': '3', 'module': '订单管理', 'description': '订单表', 'type': 'ILF', 'complexity': '中', 'fp': 10},
            {'id': '4', 'module': '支付', 'description': '支付宝接口', 'type': 'EIF', 'complexity': '低', 'fp': 5},
            {'id': '5', 'module': '用户管理', 'description': '用户注册', 'type': 'EI', 'complexity': '低', 'fp': 3},
            {'id': '6', 'module': '订单管理', 'description': '订单查询', 'type': 'EQ', 'complexity': '中', 'fp': 4},
        ]
    }
    
    excel_data = generate_excel(test_data, '电商系统')
    print(f"Excel 文件已生成，大小: {len(excel_data)} 字节")
    
    # 保存测试文件
    with open('/tmp/test_fp.xlsx', 'wb') as f:
        f.write(excel_data)
    print("测试文件已保存到 /tmp/test_fp.xlsx")
